import openpyxl as px
import numpy as np
import os

def load_radiomics(sheet):
    data_real = []
    data_imag = []
    none_features = []
    for i, row in enumerate(sheet.iter_rows()):
        if i == 0:
            continue

        temp_real = []
        temp_imag = []
        for j, col in enumerate(row):
            col_val = col.internal_value
            #print(type(col_val))
            if isinstance(col_val, str):
                complex_val = complex(col_val.replace('i','j').replace(' ',''))
                #if not isinstance(complex_val.real, float):
                #    print(type(complex_val.real))

                if np.isnan(complex_val.real):
                    #print(col_val)
                    none_features.append(j)
                temp_real.append(complex_val.real)
                temp_imag.append(complex_val.imag)
            else:
                try:
                    temp_real.append(complex(col_val).real)
                    temp_imag.append(complex(col_val).imag)
                except:
                    #print(i,j,col_val)
                    if j not in none_features:
                        none_features.append(j)
                    #print(type(col_val))
                    temp_real.append('None')
                    temp_imag.append('None')
        data_real.append(temp_real)
        data_imag.append(temp_imag)
        
    num_features = len(data_real[0])
    feature_ids = []
    data_raw = []
    none_features = none_features + [0, 1, 2] # exclude the first column and case number
    for j in range(num_features):
        if j not in none_features:
            feature_ids.append(j)
            temp = [row[j] for row in data_real]
            data_raw.append(temp)
    
    radiomics_real_np = np.asarray(data_raw, dtype='float').T#[:,1:] # exclude "survival month"

    # already remove features with nan before.
    print('data shape: {}'.format(radiomics_real_np.shape))
    print("feature_ids: {}".format(feature_ids))
    
    radiomics_feature_ids = feature_ids
    
    return radiomics_real_np, radiomics_feature_ids



def load_clinical(sheet):
    data_real = []
    data_imag = []
    none_features = []
    clinical_feats = []
    for i, row in enumerate(sheet.iter_rows()):
        if i == 0:
            for j, col in enumerate(row):
                col_val = col.internal_value
                if("R1" in col_val or col_val=="Case number" or col_val=="SurvivalMonth"):
                    none_features.append(j)
                clinical_feats.append(col_val)


            continue

        temp_real = []
        temp_imag = []
        for j, col in enumerate(row):
            col_val = col.internal_value
            #print(type(col_val))
            if isinstance(col_val, str):
                complex_val = complex(col_val.replace('i','j').replace(' ',''))
                #if not isinstance(complex_val.real, float):
                #    print(type(complex_val.real))

                if np.isnan(complex_val.real):
                    #print(col_val)
                    none_features.append(j)
                temp_real.append(complex_val.real)
                temp_imag.append(complex_val.imag)
            else:
                try:
                    temp_real.append(complex(col_val).real)
                    temp_imag.append(complex(col_val).imag)
                except:
                    #print(i,j,col_val)
                    if j not in none_features:
                        none_features.append(j)
                    #print(type(col_val))
                    temp_real.append('None')
                    temp_imag.append('None')
        data_real.append(temp_real)
        data_imag.append(temp_imag)
        
    num_features = len(data_real[0])
    data_raw = []
    for j in range(num_features):
        if j not in none_features:
            temp = [row[j] for row in data_real]
            data_raw.append(temp)

    clinical_real_np = np.asarray(data_raw, dtype='float').T # exclude the first column "survival month"

    # already remove features with nan before.
    print('data shape: {}'.format(clinical_real_np.shape))

    # clinical features
    select_clinical_features = []
    for i, feats in enumerate(clinical_feats):
        if i not in none_features:
            select_clinical_features.append(feats)

    print(select_clinical_features)
    
    return clinical_real_np, select_clinical_features


def load_all(sheet):
    data_real = []
    data_imag = []
    none_features = []
    all_feats = []
    for i, row in enumerate(sheet.iter_rows()):
        if i == 0:
            for j, col in enumerate(row):
                col_val = col.internal_value
                if col_val == None:
                    all_feats.append(j)
                    continue
                if("R1" in col_val or col_val=="Case number" or col_val=="SurvivalMonth"):
                    none_features.append(j)
                all_feats.append(col_val)


            continue

        temp_real = []
        temp_imag = []
        for j, col in enumerate(row):
            col_val = col.internal_value
            #print(type(col_val))
            if isinstance(col_val, str):
                complex_val = complex(col_val.replace('i','j').replace(' ',''))
                #if not isinstance(complex_val.real, float):
                #    print(type(complex_val.real))

                if np.isnan(complex_val.real):
                    #print(col_val)
                    none_features.append(j)
                temp_real.append(complex_val.real)
                temp_imag.append(complex_val.imag)
            else:
                try:
                    temp_real.append(complex(col_val).real)
                    temp_imag.append(complex(col_val).imag)
                except:
                    #print(i,j,col_val)
                    if j not in none_features:
                        none_features.append(j)
                    #print(type(col_val))
                    temp_real.append('None')
                    temp_imag.append('None')
        data_real.append(temp_real)
        data_imag.append(temp_imag)
        
    num_features = len(data_real[0])
    data_raw = []
    for j in range(num_features):
        if j not in none_features:
            temp = [row[j] for row in data_real]
            data_raw.append(temp)

    all_real_np = np.asarray(data_raw, dtype='float').T # exclude the first column "survival month"

    # already remove features with nan before.
    print('data shape: {}'.format(all_real_np.shape))

    # all features
    select_all_features = []
    for i, feats in enumerate(all_feats):
        if i not in none_features:
            select_all_features.append(feats)

    print(select_all_features)
    
    return all_real_np, select_all_features
        
    